import openpyxl

# wb = openpyxl.Workbook()
wb = openpyxl.load_workbook("transactions.xlsx")
print(wb.sheetnames)

# access
sheet = wb["Sheet1"]
print('\n')

cell = sheet["a1"]
column = sheet["a"]
print(column)
print('\n')

cells = sheet["a:c"] # range of column
print(cells)
print('\n')

sheet["a1:c3"]
print(cells)
sheet[1:3]

sheet.append([1, 2, 3])
# sheet.delete_cols

# to save in new file 
wb.save("transactions2.xlsx")





cell = sheet.cell(row=1, column=1)  # this is useful when we're iterating
# over all the rows and all the columns and you want to dynamically access
# various cells.
print('\n')

print(sheet.max_row)
print(sheet.max_column)
print('\n')

for row in range(1, sheet.max_row + 1):
    for column in range(1, sheet.max_column + 1):
        cell = sheet.cell(row, column)
        print(cell.value)


# print(cell.value)
# print(cell.row)
# print(cell.column)
# print(cell.coordinate)
# cell.value = 1 # change the value here

# wb.create_sheet("Sheet2", 0)
# wb.remove_sheet(sheet)
